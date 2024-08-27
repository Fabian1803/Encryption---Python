import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import subprocess  # Para abrir el archivo Excel en el sistema operativo

# Crear una tabla de desplazamientos
tabla_desplazamientos = {i: i - 26 for i in range(1, 26)}
tabla_desplazamientos.update({i: i for i in range(1, 26)})

def obtener_desplazamiento_ajustado(valor_c, valor_k):
    resultado = valor_c - valor_k
    if resultado > 25:
        # Ajustar el resultado usando la tabla de desplazamientos
        return tabla_desplazamientos[(resultado + 26) % 26]
    else:
        # No se necesita ajuste
        return resultado

def cifrar_vigenere_con_calculos(texto_plano, clave):
    clave = clave.upper()
    texto_plano = texto_plano.upper().replace(' ', '')  # Eliminar espacios para el cifrado
    clave_repetida = (clave * (len(texto_plano) // len(clave) + 1))[:len(texto_plano)]
    
    texto_cifrado = []
    calculos = []
    for p, k in zip(texto_plano, clave_repetida):
        if p.isalpha():  # Cifra solo caracteres alfabéticos
            p_val = ord(p) - ord('A')
            k_val = ord(k) - ord('A')
            desplazamiento = (p_val + k_val) % 26
            nuevo_char = chr(desplazamiento + ord('A'))
            texto_cifrado.append(nuevo_char)
            calculos.append((p, k, p_val, k_val, nuevo_char, desplazamiento))
    
    return ''.join(texto_cifrado), clave_repetida, calculos

def descifrar_vigenere_con_calculos(texto_cifrado, clave):
    clave = clave.upper()
    texto_cifrado = texto_cifrado.upper().replace(' ', '')  # Eliminar espacios para el descifrado
    clave_repetida = (clave * (len(texto_cifrado) // len(clave) + 1))[:len(texto_cifrado)]
    
    texto_plano = []
    calculos = []
    for c, k in zip(texto_cifrado, clave_repetida):
        if c.isalpha():  # Descifra solo caracteres alfabéticos
            c_val = ord(c) - ord('A')
            k_val = ord(k) - ord('A')
            desplazamiento = obtener_desplazamiento_ajustado(c_val, k_val)
            nuevo_char = chr((desplazamiento + ord('A')) % 26 + ord('A'))
            calculos.append((c, k, c_val, k_val, nuevo_char, desplazamiento))
            texto_plano.append(nuevo_char)
    
    return ''.join(texto_plano), clave_repetida, calculos

def guardar_en_excel_con_calculos(texto_plano, clave, clave_extendida, texto_cifrado, calculos_descifrado, nombre_archivo='resultados.xlsx'):
    """Guarda los cálculos del descifrado Vigenère en un archivo Excel con filas adicionales."""
    def formatear_calculos(calcs):
        """Formatea los cálculos en columnas."""
        return [[calc[0], calc[2], calc[1], calc[3], calc[4], calc[5]] for calc in calcs]

    calculos_descifrado = formatear_calculos(calculos_descifrado)

    # Crear un DataFrame para los cálculos
    df = pd.DataFrame(calculos_descifrado, columns=['Texto_Cifrado', 'Valor_C', 'Clave', 'Valor_K', 'Texto_Descifrado', 'Desplazamiento'])
    
    # Guardar el DataFrame en un archivo Excel temporal
    df.to_excel(nombre_archivo, index=False, engine='openpyxl')

    # Cargar el archivo Excel y modificar las celdas
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    # Escribir datos en celdas específicas
    ws['H2'] = "Texto Plano:"
    ws['H3'] = "Clave:"
    ws['H4'] = "Clave Extendida:"
    ws['H5'] = "Texto Cifrado:"
    ws['I2'] = texto_plano
    ws['I3'] = clave
    ws['I4'] = clave_extendida
    ws['I5'] = texto_cifrado

    # Centrar el texto en las celdas
    for row in ws.iter_rows(min_row=2, max_col=9, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Añadir filas en blanco arriba de los datos
    num_filas_en_blanco = 1  # Número de filas a insertar antes de los datos
    ws.insert_rows(1, num_filas_en_blanco)
    
    # Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardar los cambios
    wb.save(nombre_archivo)

    # Abrir el archivo Excel
    if os.name == 'nt':  # Para sistemas Windows
        os.startfile(nombre_archivo)
    elif os.name == 'posix':  # Para sistemas Unix (Linux/macOS)
        subprocess.run(['open', nombre_archivo], check=True)

class Vigenere(tk.Frame):
    def __init__(self, master, switch_frame_callback):
        super().__init__(master)
        self.switch_frame_callback = switch_frame_callback
        
        label = tk.Label(self, text="Este es el Frame Vigenere")
        label.pack(pady=5)
        
        # Campo de entrada de texto para cifrar
        tk.Label(self, text="Texto para Cifrar:").pack(pady=5)
        self.texto_cifrar = tk.Entry(self, width=30)
        self.texto_cifrar.pack(pady=5)
        
        # Campo de entrada de clave
        tk.Label(self, text="Clave: Ejem: ABCD").pack(pady=5)
        self.clave = tk.Entry(self, width=30)
        self.clave.pack(pady=5)

        # Botón para cifrar
        button_cifrar = tk.Button(self, text="Cifrar", command=self.cifrar_texto)
        button_cifrar.pack(pady=5)

        # Área de texto para mostrar el resultado del cifrado
        self.resultado_cifrado = tk.Text(self, height=2, width=30)
        self.resultado_cifrado.pack(pady=5)
        self.resultado_cifrado.config(state=tk.DISABLED)  # Inicialmente deshabilitado para solo lectura
        
        # Botón para guardar en Excel
        button_exel = tk.Button(self, text="Guardar en Excel", command=self.guardar_en_excel)
        button_exel.pack(pady=5)

        # Botón para regresar al Frame Menu
        button = tk.Button(self, text="Regresar al Frame Menu", command=self.switch_frame_callback)
        button.pack(pady=5)

    def cifrar_texto(self):
        texto = self.texto_cifrar.get()
        clave = self.clave.get()
        
        if texto and clave:
            texto_cifrado, clave_extendida, _ = cifrar_vigenere_con_calculos(texto, clave)
            
            # Mostrar resultados en el área de texto
            self.resultado_cifrado.config(state=tk.NORMAL)
            self.resultado_cifrado.delete(1.0, tk.END)
            self.resultado_cifrado.insert(tk.END, f"Texto Cifrado: {texto_cifrado}")
            self.resultado_cifrado.config(state=tk.DISABLED)

    def guardar_en_excel(self):
        texto = self.texto_cifrar.get()
        clave = self.clave.get()
        
        if texto and clave:
            texto_cifrado, clave_extendida, calculos_cifrado = cifrar_vigenere_con_calculos(texto, clave)
            _, _, calculos_descifrado = descifrar_vigenere_con_calculos(texto_cifrado, clave)
            
            # Guardar en Excel
            nombre_archivo = 'resultados_vigenere.xlsx'
            guardar_en_excel_con_calculos(texto, clave, clave_extendida, texto_cifrado, calculos_descifrado, nombre_archivo)
        else:
            tk.messagebox.showwarning("Entrada incompleta", "Por favor, ingrese el texto y la clave.")

