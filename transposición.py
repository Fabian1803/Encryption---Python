import os
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class Transposición(tk.Frame):
    def __init__(self, master, switch_frame_callback):
        super().__init__(master)
        self.switch_frame_callback = switch_frame_callback
        
        label = tk.Label(self, text="Este es el Frame Transposición")
        label.pack(pady=5)

        # Campo de entrada de texto para cifrar
        tk.Label(self, text="Texto para Cifrar:").pack(pady=5)
        self.texto_cifrar = tk.Entry(self, width=30)
        self.texto_cifrar.pack(pady=5)

        # Campo de entrada para la transposición
        tk.Label(self, text="Desplazamiento para Cifrar (e.g., 2,1,3):").pack(pady=5)
        self.desplazamiento = tk.Entry(self, width=30)
        self.desplazamiento.pack(pady=5)
        
        # Botón para cifrar
        button_cifrar = tk.Button(self, text="Cifrar", command=self.cifrar_texto)
        button_cifrar.pack(pady=5)

        # Área de texto para mostrar el resultado del cifrado
        self.resultado_cifrar = tk.Text(self, height=2, width=30)
        self.resultado_cifrar.pack(pady=5)
        self.resultado_cifrar.config(state=tk.DISABLED)  # Inicialmente deshabilitado para solo lectura

        # Botón para guardar en Excel
        button_exel = tk.Button(self, text="Guardar en Excel", command=self.guardar_en_excel)
        button_exel.pack(pady=5)
        
        button = tk.Button(self, text="Regresar al Frame Menu", command=self.switch_frame_callback)
        button.pack(pady=5)
    
    def transposition_cipher_block(self, block, permutation):
        """Aplica el cifrado por transposición a un solo bloque sin usar bucles for."""
        encrypted_block = ''.join(block[p - 1] for p in permutation)
        return encrypted_block

    def inverse_permutation(self, permutation):
        """Devuelve la permutación inversa dada una permutación."""
        inverse_perm = [0] * len(permutation)
        for i, p in enumerate(permutation):
            inverse_perm[p - 1] = i + 1
        return inverse_perm

    def transposition_decipher_block(self, block, permutation):
        """Desencripta un solo bloque usando la permutación inversa."""
        inverse_perm = self.inverse_permutation(permutation)
        decrypted_block = ''.join(block[p - 1] for p in inverse_perm)
        return decrypted_block
    
    def save_text_to_excel(self, text, permutation, encrypted_text, filename):
        """Guarda el texto en un archivo Excel con varias filas, incluyendo texto y permutación."""
        # Convertir el texto y el texto cifrado en listas de caracteres
        text_data = [list(text)]
        encrypted_data = [list(encrypted_text)]
        
        # Crear una lista de índices de acuerdo con la permutación ordenada (1, 2, 3, 4, 5 repetido)
        perm_indices = [(i % len(permutation)) + 1 for i in range(len(text))]
        
        # Crear una lista de índices de permutación para el texto cifrado
        encrypted_perm_indices = [str(p) for p in permutation]
        encrypted_perm_indices_repeated = [encrypted_perm_indices[i % len(encrypted_perm_indices)] for i in range(len(encrypted_text))]
        
        # Crear un DataFrame con las filas correspondientes
        df = pd.DataFrame([
            ['Texto plano:'],
            text_data[0],
            perm_indices,
            [''] * len(text_data[0]),  # Espacio vacío
            ['Texto encriptado:'],
            encrypted_data[0],
            encrypted_perm_indices_repeated,
            [''] * len(text_data[0]),  # Espacio vacío
            ['Texto desencriptado:'],
            encrypted_data[0],
            encrypted_perm_indices_repeated,
            [''] * len(text_data[0]),  # Espacio vacío
            text_data[0],
            perm_indices
        ])
        
        # Guardar en un archivo Excel
        df.to_excel(filename, index=False, header=False)
        
        # Formatear el archivo Excel para ajustar el formato de celdas
        wb = load_workbook(filename)
        ws = wb.active
        
        # Ajustar el formato de las celdas
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row > 2:  # Las celdas de datos en las filas 1 y 2 (texto y permutación) están bien como texto.
                    try:
                        cell.value = int(cell.value)  # Intenta convertir el valor a entero
                    except (ValueError, TypeError):
                        pass  # Si no se puede convertir, mantén el valor original
        
        # Guardar los cambios en el archivo Excel
        wb.save(filename)
        print(f"\nEl archivo Excel '{filename}' ha sido creado exitosamente.")

        # Abrir el archivo Excel
        if os.name == 'nt':  # Para Windows
            os.startfile(filename)
        elif os.name == 'posix':  # Para macOS o Linux
            os.system(f'open "{filename}"' if sys.platform == 'darwin' else f'xdg-open "{filename}"')
    
    def cifrar_texto(self):
        try:
            # Obtener el texto y la permutación desde los campos de entrada
            texto = self.texto_cifrar.get().replace(' ', '')
            permutacion = list(map(int, self.desplazamiento.get().split(',')))
            
            # Ajustar el texto al tamaño del bloque y agregar espacios si es necesario
            block_size = len(permutacion)
            num_blocks = (len(texto) + block_size - 1) // block_size
            texto_padded = texto.ljust(num_blocks * block_size, ' ')
            
            # Dividir el texto en bloques
            bloques = [texto_padded[i * block_size:(i + 1) * block_size] for i in range(num_blocks)]
            
            # Encriptar cada bloque y mostrar el resultado
            bloques_encriptados = [self.transposition_cipher_block(bloque, permutacion) for bloque in bloques]
            
            # Combinar los bloques cifrados
            texto_encriptado = ''.join(bloques_encriptados)
            
            # Dividir el texto cifrado en bloques
            bloques_encriptados = [texto_encriptado[i * block_size:(i + 1) * block_size] for i in range(num_blocks)]
            
            # Desencriptar cada bloque y mostrar el resultado
            bloques_desencriptados = [self.transposition_decipher_block(bloque, permutacion) for bloque in bloques_encriptados]
            
            # Combinar los bloques desencriptados y eliminar espacios adicionales
            texto_desencriptado = ''.join(bloques_desencriptados).rstrip()
            
            # Mostrar el resultado en el área de texto
            self.resultado_cifrar.config(state=tk.NORMAL)
            self.resultado_cifrar.delete(1.0, tk.END)
            self.resultado_cifrar.insert(tk.END, f"Texto cifrado: {texto_encriptado}\n")
            self.resultado_cifrar.insert(tk.END, f"Texto desencriptado: {texto_desencriptado}\n")
            self.resultado_cifrar.config(state=tk.DISABLED)
        
        except Exception as e:
            messagebox.showerror("Error", f"Hubo un problema: {e}")

    def guardar_en_excel(self):
        try:
            texto = self.texto_cifrar.get().replace(' ', '')
            permutacion = list(map(int, self.desplazamiento.get().split(',')))
            
            # Ajustar el texto al tamaño del bloque y agregar espacios si es necesario
            block_size = len(permutacion)
            num_blocks = (len(texto) + block_size - 1) // block_size
            texto_padded = texto.ljust(num_blocks * block_size, ' ')
            
            # Dividir el texto en bloques
            bloques = [texto_padded[i * block_size:(i + 1) * block_size] for i in range(num_blocks)]
            
            # Encriptar cada bloque
            bloques_encriptados = [self.transposition_cipher_block(bloque, permutacion) for bloque in bloques]
            texto_encriptado = ''.join(bloques_encriptados)
            
            # Guardar el texto en un archivo Excel
            filename = os.path.join(os.getcwd(), 'resultados.xlsx')
            self.save_text_to_excel(texto, permutacion, texto_encriptado, filename)
        
        except Exception as e:
            messagebox.showerror("Error", f"Hubo un problema: {e}")
